"""Professional Excel estimate workbook generation.

The workbook is a presentation artifact only. Every numeric value comes from the
validated deterministic pricing engine; this module performs no estimating and
refuses to render unreconciled or component-mismatched totals.
"""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_MONEY_FORMAT = '$#,##0.00;[Red]-$#,##0.00'
_QUANTITY_FORMAT = '#,##0.0000'
_HEADER_FILL = PatternFill("solid", fgColor="17365D")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")


def _decimal(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, bool):
        raise ValueError("boolean is not a valid estimate number")
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError) as exc:
        raise ValueError("invalid deterministic estimate number") from exc


def _money(value: Any) -> float:
    return float(_decimal(value).quantize(Decimal("0.01")))


def _text_list(value: Any) -> list[str]:
    if not isinstance(value, list):
        return []
    return [str(item).strip() for item in value if str(item).strip()]


def _style_header(row) -> None:  # noqa: ANN001 - openpyxl row cells
    for cell in row:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _fit_columns(ws, *, maximum: int = 42) -> None:  # noqa: ANN001
    for column_cells in ws.columns:
        width = min(max(len(str(c.value or "")) for c in column_cells) + 2, maximum)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max(width, 10)


def _validate_totals(lines: list[dict[str, Any]], rollup: dict[str, Any]) -> None:
    if not rollup.get("reconciled"):
        raise ValueError("estimate rollup is not reconciled")
    component_fields = (
        "labor_cost",
        "material_cost",
        "equipment_cost",
        "subcontract_cost",
        "other_direct_cost",
    )
    direct_sum = Decimal("0")
    for line in lines:
        component_total = sum((_decimal(line.get(key)) for key in component_fields), Decimal("0"))
        declared = _decimal(line.get("direct_cost_total"))
        if component_total.quantize(Decimal("0.01")) != declared.quantize(Decimal("0.01")):
            raise ValueError("line-item components do not reconcile to direct total")
        direct_sum += declared
    expected = _decimal((rollup.get("totals") or {}).get("direct_cost_subtotal"))
    if direct_sum.quantize(Decimal("0.01")) != expected.quantize(Decimal("0.01")):
        raise ValueError("line-item totals do not reconcile to estimate rollup")


def build_estimate_workbook(
    *,
    project: dict[str, Any],
    estimate: dict[str, Any],
    version: dict[str, Any],
    lines: list[dict[str, Any]],
    rollup: dict[str, Any],
    review_events: Iterable[dict[str, Any]] = (),
) -> bytes:
    """Return a validated XLSX workbook for one deterministic estimate version."""

    _validate_totals(lines, rollup)
    totals = rollup.get("totals") or {}

    wb = Workbook()
    summary = wb.active
    if summary is None:  # defensive for type checkers; Workbook always creates one sheet
        raise ValueError("workbook did not create an active worksheet")
    summary.title = "Project Summary"
    summary.sheet_view.showGridLines = False
    summary["A1"] = "MOBI ESTIMATES"
    summary["A1"].font = Font(size=18, bold=True, color="17365D")
    summary["A3"] = "Project"
    summary["B3"] = project.get("name") or ""
    summary["A4"] = "Estimate"
    summary["B4"] = estimate.get("name") or ""
    summary["A5"] = "Version"
    summary["B5"] = version.get("version_number")
    summary["A6"] = "Status"
    summary["B6"] = version.get("status") or ""
    summary["A7"] = "Pricing date"
    summary["B7"] = version.get("pricing_date") or ""
    summary["A8"] = "Currency"
    summary["B8"] = version.get("currency") or "USD"

    summary["A10"] = "Cost Summary"
    summary["A10"].fill = _SECTION_FILL
    summary["A10"].font = Font(bold=True)
    summary.merge_cells("A10:B10")
    summary_rows = [
        ("Labor", "labor_cost"),
        ("Materials", "material_cost"),
        ("Equipment", "equipment_cost"),
        ("Subcontractors", "subcontract_cost"),
        ("Other direct", "other_direct_cost"),
        ("Direct subtotal", "direct_cost_subtotal"),
        ("Indirect costs", "indirect_costs"),
        ("Tax", "tax"),
        ("Overhead", "overhead"),
        ("Profit", "profit"),
        ("Contingency", "contingency"),
        ("Bond", "bond"),
        ("Insurance", "insurance"),
        ("Discounts", "discounts"),
        ("Other adjustments", "other_adjustments"),
        ("FINAL TOTAL", "final_sell_price"),
    ]
    for row_index, (label, key) in enumerate(summary_rows, start=11):
        summary.cell(row_index, 1, label)
        summary.cell(row_index, 2, _money(totals.get(key)))
        summary.cell(row_index, 2).number_format = _MONEY_FORMAT
        if label == "FINAL TOTAL":
            for cell in summary[row_index]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = _HEADER_FILL
    summary["A28"] = "Deterministic verification"
    summary["B28"] = "PASS — line components and direct subtotal reconciled before export"
    summary.column_dimensions["A"].width = 25
    summary.column_dimensions["B"].width = 62

    detail = wb.create_sheet("Detailed Line Items")
    detail_headers = [
        "Trade", "Category", "Description", "Location", "Quantity", "Unit",
        "Labor", "Materials", "Equipment", "Subcontractors", "Other Direct",
        "Direct Total", "Evidence / Source References", "Status",
    ]
    detail.append(detail_headers)
    _style_header(detail[1])
    for row_index, line in enumerate(lines, start=2):
        evidence = line.get("evidence") if isinstance(line.get("evidence"), list) else []
        refs = []
        for row in evidence:
            if not isinstance(row, dict):
                continue
            parts = [row.get("source_document") or row.get("document_name"), row.get("sheet_number"), row.get("page_number")]
            ref = " / ".join(str(part) for part in parts if part not in (None, ""))
            quote = row.get("extracted_text_quote") or row.get("quote")
            refs.append(f"{ref}: {quote}" if ref and quote else ref or str(quote or ""))
        detail.append([
            line.get("trade_code") or "",
            line.get("category_code") or "",
            line.get("description") or "",
            line.get("location") or "",
            float(_decimal(line.get("quantity"))),
            line.get("unit") or "",
            _money(line.get("labor_cost")),
            _money(line.get("material_cost")),
            _money(line.get("equipment_cost")),
            _money(line.get("subcontract_cost")),
            _money(line.get("other_direct_cost")),
            f"=SUM(G{row_index}:K{row_index})",
            "\n".join(ref for ref in refs if ref),
            line.get("status") or "",
        ])
        detail.cell(row_index, 5).number_format = _QUANTITY_FORMAT
        for column in range(7, 13):
            detail.cell(row_index, column).number_format = _MONEY_FORMAT
        detail.cell(row_index, 13).alignment = Alignment(wrap_text=True, vertical="top")
    total_row = len(lines) + 2
    detail.cell(total_row, 11, "TOTAL")
    detail.cell(total_row, 11).font = Font(bold=True)
    detail.cell(total_row, 12, f"=SUM(L2:L{max(total_row - 1, 2)})")
    detail.cell(total_row, 12).number_format = _MONEY_FORMAT
    detail.cell(total_row, 12).font = Font(bold=True)
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = f"A1:N{max(total_row - 1, 1)}"
    _fit_columns(detail)

    for title, values in (
        ("Assumptions", _text_list(version.get("assumptions"))),
        ("Exclusions", _text_list(version.get("exclusions"))),
        ("RFIs & Clarifications", _text_list(version.get("clarifications"))),
        ("Allowances", _text_list((version.get("config") or {}).get("allowances") if isinstance(version.get("config"), dict) else [])),
        ("Alternates", _text_list((version.get("config") or {}).get("alternates") if isinstance(version.get("config"), dict) else [])),
    ):
        ws = wb.create_sheet(title)
        ws.append([title])
        _style_header(ws[1])
        if values:
            for number, value in enumerate(values, start=1):
                ws.append([number, value])
        else:
            ws.append(["None recorded"])
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 100
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    revision = wb.create_sheet("Revision History")
    revision.append(["Date", "Action", "Previous", "New", "Reviewer", "Notes"])
    _style_header(revision[1])
    events = list(review_events)
    if events:
        for event in events:
            revision.append([
                event.get("created_at") or "",
                event.get("action") or "",
                event.get("previous_state") or "",
                event.get("new_state") or "",
                event.get("reviewer_id") or "",
                event.get("notes") or "",
            ])
    else:
        revision.append(["", "No review events recorded", "", "", "", ""])
    _fit_columns(revision, maximum=70)

    output = BytesIO()
    wb.save(output)
    payload = output.getvalue()

    # Re-open the archive before returning it. This catches corrupt/partial XLSX
    # serialization independently of the caller or HTTP response layer.
    check = load_workbook(BytesIO(payload), data_only=False, read_only=True)
    if "Project Summary" not in check.sheetnames or "Detailed Line Items" not in check.sheetnames:
        raise ValueError("generated workbook is missing required sheets")
    check.close()
    return payload
