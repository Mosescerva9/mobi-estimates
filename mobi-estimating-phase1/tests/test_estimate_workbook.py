from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.pricing.workbook import build_estimate_workbook


def _line(**overrides):
    line = {
        "trade_code": "painting",
        "category_code": "walls",
        "description": "Prime and paint walls",
        "location": "Level 1",
        "quantity": "100",
        "unit": "SF",
        "labor_cost": "100.00",
        "material_cost": "50.00",
        "equipment_cost": "0.00",
        "subcontract_cost": "0.00",
        "other_direct_cost": "0.00",
        "direct_cost_total": "150.00",
        "status": "priced",
        "evidence": [{
            "source_document": "plans.pdf",
            "sheet_number": "A101",
            "page_number": 3,
            "extracted_text_quote": "PAINT WALLS",
        }],
    }
    line.update(overrides)
    return line


def _rollup(**total_overrides):
    totals = {
        "labor_cost": "100.00",
        "material_cost": "50.00",
        "equipment_cost": "0.00",
        "subcontract_cost": "0.00",
        "other_direct_cost": "0.00",
        "direct_cost_subtotal": "150.00",
        "indirect_costs": "10.00",
        "tax": "5.00",
        "overhead": "10.00",
        "profit": "20.00",
        "contingency": "0.00",
        "bond": "0.00",
        "insurance": "0.00",
        "discounts": "0.00",
        "other_adjustments": "0.00",
        "final_sell_price": "195.00",
    }
    totals.update(total_overrides)
    return {"totals": totals, "reconciled": True}


def test_professional_workbook_has_required_sheets_formulas_and_sources():
    payload = build_estimate_workbook(
        project={"name": "Golden Path Test"},
        estimate={"name": "Base Bid"},
        version={
            "version_number": 2,
            "status": "approved",
            "pricing_date": "2026-07-21",
            "currency": "USD",
            "assumptions": ["Normal working hours"],
            "exclusions": ["Hazardous materials"],
            "clarifications": ["Confirm finish color"],
            "config": {"allowances": ["Permit allowance"], "alternates": ["Alternate 1"]},
        },
        lines=[_line()],
        rollup=_rollup(),
        review_events=[{
            "created_at": "2026-07-21T12:00:00Z",
            "action": "approve",
            "previous_state": "priced",
            "new_state": "approved",
            "reviewer_id": "reviewer-1",
            "notes": "QA complete",
        }],
    )

    workbook = load_workbook(BytesIO(payload), data_only=False)
    assert workbook.sheetnames == [
        "Project Summary",
        "Detailed Line Items",
        "Assumptions",
        "Exclusions",
        "RFIs & Clarifications",
        "Allowances",
        "Alternates",
        "Revision History",
    ]
    detail = workbook["Detailed Line Items"]
    assert detail["L2"].value == "=SUM(G2:K2)"
    assert detail["M2"].value == "plans.pdf / A101 / 3: PAINT WALLS"
    assert workbook["Project Summary"]["B26"].value == 195.0
    assert workbook["Revision History"]["B2"].value == "approve"


def test_workbook_fails_closed_on_line_component_mismatch():
    with pytest.raises(ValueError, match="components do not reconcile"):
        build_estimate_workbook(
            project={"name": "Mismatch"},
            estimate={"name": "Estimate"},
            version={"version_number": 1},
            lines=[_line(direct_cost_total="999.00")],
            rollup=_rollup(direct_cost_subtotal="999.00"),
        )


def test_workbook_fails_closed_on_unreconciled_rollup():
    with pytest.raises(ValueError, match="not reconciled"):
        build_estimate_workbook(
            project={"name": "Mismatch"},
            estimate={"name": "Estimate"},
            version={"version_number": 1},
            lines=[_line()],
            rollup={**_rollup(), "reconciled": False},
        )
